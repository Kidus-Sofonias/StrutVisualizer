"""Check Excel reference values for 4.6."""
import sys
sys.path.insert(0, '.')
import openpyxl

wb = openpyxl.load_workbook('../Stability Behaviour (Bahru Sarbet).xlsx', data_only=True, read_only=True)
print("Sheets:", wb.sheetnames)

# Check section 4.6
for name in wb.sheetnames:
    if '4.6' in name or 'overturn' in name.lower() or 'over' in name.lower():
        ws = wb[name]
        print(f"\n=== {name} ===")
        for row in ws.iter_rows(min_row=1, max_row=40, values_only=False):
            vals = [(c.value, c.coordinate) for c in row if c.value is not None]
            if vals:
                print(vals)

# Also check section 3.4 sheet for the q-factor display
for name in wb.sheetnames:
    if '3.4' in name or 'behavioral' in name.lower() or 'behaviour' in name.lower():
        ws = wb[name]
        print(f"\n=== {name} ===")
        for row in ws.iter_rows(min_row=1, max_row=30, values_only=False):
            vals = [(c.value, c.coordinate) for c in row if c.value is not None]
            if vals:
                print(vals)
