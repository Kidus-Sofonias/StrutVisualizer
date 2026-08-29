"""
Comprehensive comparison: Our computed values vs Excel workbook
for sections 3.2.5, 3.2.6, 3.2.7 and check for any blank values.
"""
import sys, os
sys.path.insert(0, os.path.dirname(__file__))

from openpyxl import load_workbook
from pathlib import Path
import pickle

# ── 1. Extract Excel reference values ──────────────────────────────────
excel_path = Path(__file__).parent.parent / "Stability_Behaviour_Calculation_for BAHRU SARBET MODEL 1-3 Final =1 (1).xlsx"
if not excel_path.exists():
    # Try finding it
    import glob
    matches = glob.glob(str(Path(__file__).parent.parent / "*.xlsx"))
    if matches:
        excel_path = Path(matches[0])
    else:
        print("ERROR: Excel file not found")
        sys.exit(1)

print(f"Loading Excel: {excel_path.name}")
wb = load_workbook(str(excel_path), data_only=True)
print(f"Sheets: {wb.sheetnames}")

# Find the sheet with 3.2 data
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    # Look for 3.2.5 or 3.2.6 or 3.2.7 headers
    for row in ws.iter_rows(min_row=1, max_row=min(10, ws.max_row), values_only=False):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                v = cell.value.strip()
                if '3.2.5' in v or '3.2.6' in v or '3.2.7' in v or 'torsional radius' in v.lower() or 'stiffness' in v.lower() or 'rx' in v.lower() or 'ls' in v.lower():
                    print(f"\nFound in sheet '{sheet_name}', cell {cell.coordinate}: {v[:80]}")

# ── Search for relevant sheets ──
print("\n\n=== SEARCHING ALL SHEETS FOR 3.2 DATA ===")
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    found_keywords = set()
    for row in ws.iter_rows(min_row=1, max_row=min(50, ws.max_row), values_only=True):
        for val in row:
            if val and isinstance(val, str):
                v = val.strip().lower()
                if '3.2.5' in v: found_keywords.add('3.2.5')
                if '3.2.6' in v: found_keywords.add('3.2.6')  
                if '3.2.7' in v: found_keywords.add('3.2.7')
                if 'ls' == v or 'floor radius' in v: found_keywords.add('ls/floor_radius')
                if 'rx' == v or 'ry' == v: found_keywords.add('rx/ry')
                if 'stiffness' in v: found_keywords.add('stiffness')
                if 'kx' == v or 'ky' == v: found_keywords.add('kx/ky')
    if found_keywords:
        print(f"\nSheet '{sheet_name}' contains: {found_keywords}")
        # Print first 80 rows to find the data
        print("  First rows with data:")
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=min(80, ws.max_row), values_only=True), 1):
            vals = [str(v)[:20] if v is not None else '' for v in row[:15]]
            if any(vals):
                print(f"    Row {i:3d}: {' | '.join(vals)}")
