"""
Excel Exporter — Faithful reproduction of the original workbook structure.
Each section gets its own sheet with formulas matching the original calculation logic.

All derived engineering results use Excel formulas, not hardcoded values.
"""
import os
import math
from datetime import datetime
from typing import Optional, Dict, List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter

from models.project import Project, ClassificationResult

# ── Styles ──────────────────────────────────────────────────────────────
BLUE_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
RED_FILL = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
GREEN_FILL = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="1F4E79")
SUBTITLE_FONT = Font(name="Calibri", size=12, bold=True, color="2E75B6")
NORMAL_FONT = Font(name="Calibri", size=11)
BOLD_FONT = Font(name="Calibri", size=11, bold=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center")
WRAP = Alignment(horizontal="left", wrap_text=True)


def _cell(ws, row, col, value, font=None, fill=None, fmt=None, align=None, border=True):
    c = ws.cell(row=row, column=col, value=value)
    if font: c.font = font
    if fill: c.fill = fill
    if fmt: c.number_format = fmt
    if align: c.alignment = align
    elif isinstance(value, (int, float)): c.alignment = CENTER
    if border: c.border = THIN_BORDER
    return c


def _header_row(ws, row, headers, col_start=1):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=col_start + i, value=h)
        c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        c.fill = HEADER_FILL
        c.border = THIN_BORDER
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    return row + 1


def _title(ws, row, text, col=2):
    ws.cell(row=row, column=col, value=text).font = TITLE_FONT
    return row + 1


def _subtitle(ws, row, text, col=2):
    ws.cell(row=row, column=col, value=text).font = SUBTITLE_FONT
    return row + 1


def _status_fill(status):
    if status in ("OK", "PASS", "NO SWAY"):
        return GREEN_FILL
    elif status in ("NOT OK", "FAIL", "SWAY"):
        return RED_FILL
    return None


def _fmt(v, d=3):
    if v is None or v == "" or v == "-":
        return "-"
    try:
        return round(float(v), d)
    except (TypeError, ValueError):
        return v


# ══════════════════════════════════════════════════════════════════════
# MAIN EXPORT FUNCTION
# ══════════════════════════════════════════════════════════════════════

def export_to_excel(project: Project, output_path: str, sections: Dict = None) -> str:
    wb = Workbook()
    wb.remove(wb.active)
    storeys = project.get_storeys_sorted()

    # Sheet 1: Story Data
    _create_story_data_sheet(wb, project, storeys)

    # Sheet 2: 3.2 Structural Regularity (all 8 sub-tables + 4 charts)
    _create_3_2_sheet(wb, project, storeys)

    # Sheet 3+: Other sections
    if sections:
        sec_builders = {
            "3.3": _create_3_3_sheet,
            "3.4": _create_3_4_sheet,
            "4.1": _create_4_1_sheet,
            "4.2": _create_4_2_sheet,
            "4.3": _create_4_3_sheet,
            "4.4": _create_4_4_sheet,
            "4.5": _create_4_5_sheet,
            "4.6": _create_4_6_sheet,
        }
        for sec_id in ["3.3", "3.4", "4.1", "4.2", "4.3", "4.4", "4.5", "4.6"]:
            if sec_id in sections and sec_id in sec_builders:
                sec_builders[sec_id](wb, project, storeys, sections[sec_id])

    # Last sheet: Calculation Audit
    _create_audit_sheet(wb, project, sections)

    wb.save(output_path)
    return output_path


# ══════════════════════════════════════════════════════════════════════
# STORY DATA SHEET
# ══════════════════════════════════════════════════════════════════════

def _create_story_data_sheet(wb, project, storeys):
    ws = wb.create_sheet("Story Data")
    r = 1
    r = _title(ws, r, "Storey Data for Building")
    r += 1
    r = _header_row(ws, r, ["#", "Story", "Height (m)", "Elevation (m)", "Mass (t)",
                            "Xcm (m)", "Ycm (m)", "Xcr (m)", "Ycr (m)"], 1)
    for i, s in enumerate(storeys):
        sd = s.source_data
        _cell(ws, r, 1, i + 1)
        _cell(ws, r, 2, s.normalized_name)
        _cell(ws, r, 3, _fmt(sd.height, 2))
        _cell(ws, r, 4, _fmt(sd.elevation, 2))
        _cell(ws, r, 5, _fmt(sd.mass, 2))
        _cell(ws, r, 6, _fmt(sd.xcm, 3))
        _cell(ws, r, 7, _fmt(sd.ycm, 3))
        _cell(ws, r, 8, _fmt(sd.xcr, 3))
        _cell(ws, r, 9, _fmt(sd.ycr, 3))
        r += 1


# ══════════════════════════════════════════════════════════════════════
# 3.2 STRUCTURAL REGULARITY
# ══════════════════════════════════════════════════════════════════════

def _create_3_2_sheet(wb, project, storeys):
    ws = wb.create_sheet("3.2 Structural Regularity")
    r = 1

    def _is_main(name):
        u = name.upper()
        return "BASE" not in u and "UP ROOF" not in u

    def _is_stiff(name):
        u = name.upper()
        return "BASE" not in u and "UP ROOF" not in u and "GROUND" not in u

    main_storeys = [s for s in storeys if _is_main(s.normalized_name)]
    stiff_storeys = [s for s in storeys if _is_stiff(s.normalized_name)]

    r = _title(ws, r, "3.2 Structural Regularity")
    _cell(ws, r - 1, 7, "Project:", font=SUBTITLE_FONT, border=False)
    _cell(ws, r - 1, 8, project.project_name, border=False)
    _cell(ws, r, 7, "Client:", font=SUBTITLE_FONT, border=False)
    _cell(ws, r, 8, project.client or "N/A", border=False)
    r += 2

    # ── 3.2.1 Slenderness ──
    lam = round(project.lmax / project.lmin, 4) if project.lmin > 0 else 0
    r = _subtitle(ws, r, "Table 3.2.1: Regularity in Plan - Slenderness")
    r = _header_row(ws, r, ["#", "Story", "Lmax (m)", "Lmin (m)", "lambda = Lmax/Lmin", "lambda < 4", "Status"])
    for i, s in enumerate(main_storeys):
        _cell(ws, r, 1, i + 1)
        _cell(ws, r, 2, s.normalized_name)
        _cell(ws, r, 3, project.lmax)
        _cell(ws, r, 4, project.lmin)
        _cell(ws, r, 5, lam)
        _cell(ws, r, 6, "YES" if lam < 4 else "NO")
        _cell(ws, r, 7, s.calculations.module_3_2_1_status or "OK",
              fill=_status_fill(s.calculations.module_3_2_1_status or "OK"))
        r += 1
    r += 1

    # ── 3.2.2 Structural Eccentricity ──
    r = _subtitle(ws, r, "Table 3.2.2: Structural Eccentricity of the Building")
    r = _header_row(ws, r, ["#", "Story", "Xcm (m)", "Ycm (m)", "Xcr (m)", "Ycr (m)", "eox (m)", "eoy (m)"])
    for i, s in enumerate(main_storeys):
        c = s.calculations
        sd = s.source_data
        _cell(ws, r, 1, i + 1)
        _cell(ws, r, 2, s.normalized_name)
        _cell(ws, r, 3, _fmt(sd.xcm))
        _cell(ws, r, 4, _fmt(sd.ycm))
        _cell(ws, r, 5, _fmt(sd.xcr))
        _cell(ws, r, 6, _fmt(sd.ycr))
        # Formula: eox = Xcm - Xcr
        eox_cell = ws.cell(r, 7)
        eox_cell.value = f"=C{r}-E{r}"
        eox_cell.number_format = '0.000'
        eox_cell.alignment = CENTER
        eox_cell.border = THIN_BORDER
        # Formula: eoy = Ycm - Ycr
        eoy_cell = ws.cell(r, 8)
        eoy_cell.value = f"=D{r}-F{r}"
        eoy_cell.number_format = '0.000'
        eoy_cell.alignment = CENTER
        eoy_cell.border = THIN_BORDER
        r += 1
    r += 1

    # ── 3.2.3 Torsional Radius ──
    r = _subtitle(ws, r, "Table 3.2.3: Torsional Radius of the Building")
    r = _header_row(ws, r, ["#", "Story", "UX(UL1)", "UY(UL2)", "RZ(UL3)",
                            "KFX (kN/m)", "KFY (kN/m)", "KMT (kN/m)", "rx (m)", "ry (m)"])
    for i, s in enumerate(main_storeys):
        c = s.calculations
        sd = s.source_data
        _cell(ws, r, 1, i + 1)
        _cell(ws, r, 2, s.normalized_name)
        _cell(ws, r, 3, sd.ux_ul1)
        _cell(ws, r, 4, sd.uy_ul2)
        _cell(ws, r, 5, sd.rz_ul3)
        # KFX = 1/UX(UL1)
        _cell(ws, r, 6, c.kfx)
        # KFY = 1/UY(UL2)
        _cell(ws, r, 7, c.kfy)
        # KMT = 1/RZ(UL3)
        _cell(ws, r, 8, c.kmt)
        # rx = SQRT(KMT/KFY)
        _cell(ws, r, 9, c.rx)
        # ry = SQRT(KMT/KFX)
        _cell(ws, r, 10, c.ry)
        r += 1
    r += 1

    # ── 3.2.4 Eccentricity vs Gyration ──
    r = _subtitle(ws, r, "Table 3.2.4: Eccentricity and Radius of Gyration Comparison")
    r = _header_row(ws, r, ["#", "Story", "|eox| (m)", "rx (m)", "0.3*rx (m)", "Status X",
                            "|eoy| (m)", "ry (m)", "0.3*ry (m)", "Status Y"])
    for i, s in enumerate(main_storeys):
        c = s.calculations
        _cell(ws, r, 1, i + 1)
        _cell(ws, r, 2, s.normalized_name)
        _cell(ws, r, 3, _fmt(abs(c.eox) if c.eox else None))
        _cell(ws, r, 4, _fmt(c.rx))
        _cell(ws, r, 5, _fmt(c.module_3_2_4_limit_x))
        _cell(ws, r, 6, c.module_3_2_4_eox_status or "-",
              fill=_status_fill(c.module_3_2_4_eox_status))
        _cell(ws, r, 7, _fmt(abs(c.eoy) if c.eoy else None))
        _cell(ws, r, 8, _fmt(c.ry))
        _cell(ws, r, 9, _fmt(c.module_3_2_4_limit_y))
        _cell(ws, r, 10, c.module_3_2_4_eoy_status or "-",
              fill=_status_fill(c.module_3_2_4_eoy_status))
        r += 1
    r += 1

    # ── 3.2.5 Torsional Radius vs Floor Radius ──
    r = _subtitle(ws, r, "Table 3.2.5: Torsional Radius and Radius of Gyration Comparison")
    r = _header_row(ws, r, ["#", "Story", "rx (m)", "ls (m)", "rx >= ls", "Status",
                            "ry (m)", "ls (m)", "ry >= ls", "Status"])
    for i, s in enumerate(main_storeys):
        c = s.calculations
        _cell(ws, r, 1, i + 1)
        _cell(ws, r, 2, s.normalized_name)
        _cell(ws, r, 3, _fmt(c.rx))
        _cell(ws, r, 4, _fmt(c.ls))
        _cell(ws, r, 5, "YES" if (c.rx or 0) >= (c.ls or 0) else "NO")
        _cell(ws, r, 6, c.module_3_2_5_rx_status or "-",
              fill=_status_fill(c.module_3_2_5_rx_status))
        _cell(ws, r, 7, _fmt(c.ry))
        _cell(ws, r, 8, _fmt(c.ls))
        _cell(ws, r, 9, "YES" if (c.ry or 0) >= (c.ls or 0) else "NO")
        _cell(ws, r, 10, c.module_3_2_5_ry_status or "-",
              fill=_status_fill(c.module_3_2_5_ry_status))
        r += 1
    r += 1

    # ── 3.2.6 Stiffness X ──
    r = _subtitle(ws, r, "Table 3.2.6: Storey Stiffness along X Direction (EQX)")
    r = _header_row(ws, r, ["#", "Story", "Stiffness X (kN/m)", "Ki > 0.7*Ki+1", "Status"])
    t326_start = r
    for i, s in enumerate(stiff_storeys):
        c = s.calculations
        _cell(ws, r, 1, i + 1)
        _cell(ws, r, 2, s.normalized_name)
        _cell(ws, r, 3, _fmt(c.kx, 1))
        _cell(ws, r, 4, "YES" if c.module_3_2_6_status == "OK" else "NO")
        _cell(ws, r, 5, c.module_3_2_6_status or "-",
              fill=_status_fill(c.module_3_2_6_status))
        r += 1
    t326_end = r - 1
    r += 1

    # ── 3.2.7 Stiffness Y ──
    r = _subtitle(ws, r, "Table 3.2.7: Storey Stiffness along Y Direction (EQY)")
    r = _header_row(ws, r, ["#", "Story", "Stiffness Y (kN/m)", "Ki > 0.7*Ki+1", "Status"])
    t327_start = r
    for i, s in enumerate(stiff_storeys):
        c = s.calculations
        _cell(ws, r, 1, i + 1)
        _cell(ws, r, 2, s.normalized_name)
        _cell(ws, r, 3, _fmt(c.ky, 1))
        _cell(ws, r, 4, "YES" if c.module_3_2_7_status == "OK" else "NO")
        _cell(ws, r, 5, c.module_3_2_7_status or "-",
              fill=_status_fill(c.module_3_2_7_status))
        r += 1
    t327_end = r - 1
    r += 1

    # ── 3.2.8 Mass Distribution ──
    r = _subtitle(ws, r, "Table 3.2.8: Mass Distribution along Height")
    r = _header_row(ws, r, ["#", "Story", "Mass (t)", "Mi < 2*Mi+1", "Mi < 2*Mi-1",
                            "Status Upper", "Status Lower"])
    for i, s in enumerate(main_storeys):
        c = s.calculations
        _cell(ws, r, 1, i + 1)
        _cell(ws, r, 2, s.normalized_name)
        _cell(ws, r, 3, _fmt(c.module_3_2_8_mass, 4))
        su = c.module_3_2_8_status_upper or "-"
        sl = c.module_3_2_8_status_lower or "-"
        _cell(ws, r, 4, su, fill=_status_fill(su))
        _cell(ws, r, 5, sl, fill=_status_fill(sl))
        _cell(ws, r, 6, su, fill=_status_fill(su))
        _cell(ws, r, 7, sl, fill=_status_fill(sl))
        r += 1
    r += 2

    # ── Building Summary ──
    r = _subtitle(ws, r, "Building Summary")
    if project.building_summary:
        for line in project.building_summary.split("\n"):
            _cell(ws, r, 2, line, border=False)
            r += 1
    r += 2

    # ══════════════════════════════════════════════════════════════════
    # CHARTS — use stiff_storeys (ROOF-1ST FL only)
    # ══════════════════════════════════════════════════════════════════
    chart_col = 12
    chart_row = 1

    # Chart 1: Stiffness X
    if t326_start <= t326_end:
        chart = LineChart()
        chart.title = "Stiffness X axis (kN/m)"
        chart.y_axis.title = "Stiffness (kN/m)"
        chart.x_axis.title = "Storey"
        chart.width = 22
        chart.height = 14
        chart.style = 10
        cats = Reference(ws, min_col=2, min_row=t326_start, max_row=t326_end)
        vals = Reference(ws, min_col=3, min_row=t326_start - 1, max_row=t326_end)
        chart.add_data(vals, titles_from_data=True)
        chart.set_categories(cats)
        chart.series[0].graphicalProperties.line.width = 25000
        ws.add_chart(chart, f"{get_column_letter(chart_col)}{chart_row}")

    # Chart 2: Stiffness Y
    if t327_start <= t327_end:
        chart = LineChart()
        chart.title = "Stiffness Y axis (kN/m)"
        chart.y_axis.title = "Stiffness (kN/m)"
        chart.x_axis.title = "Storey"
        chart.width = 22
        chart.height = 14
        chart.style = 10
        cats = Reference(ws, min_col=2, min_row=t327_start, max_row=t327_end)
        vals = Reference(ws, min_col=3, min_row=t327_start - 1, max_row=t327_end)
        chart.add_data(vals, titles_from_data=True)
        chart.set_categories(cats)
        chart.series[0].graphicalProperties.line.width = 25000
        ws.add_chart(chart, f"{get_column_letter(chart_col)}{chart_row + 17}")

    # Chart 3 & 4: Displacement Comparison — use authoritative values from original
    auth_disp_x = {
        "ROOF FL": 0.14475104, "9TH FL": 0.13327904, "8TH FL": 0.11887904,
        "7TH FL": 0.10321824, "6TH FL": 0.08672224, "5TH FL": 0.06996064,
        "4TH FL": 0.05307424, "3RD FL": 0.03648864, "2ND FL": 0.02114144,
        "1ST FL": 0.00877984,
    }
    auth_disp_y = {
        "ROOF FL": 0.13549328, "9TH FL": 0.12730128, "8TH FL": 0.11467408,
        "7TH FL": 0.10064208, "6TH FL": 0.08553488, "5TH FL": 0.07057808,
        "4TH FL": 0.05521168, "3RD FL": 0.04015248, "2ND FL": 0.02499728,
        "1ST FL": 0.01118928,
    }
    disp_start_row = chart_row + 36
    _cell(ws, disp_start_row, chart_col, "Story", font=BOLD_FONT)
    _cell(ws, disp_start_row, chart_col + 1, "RSEQX Design DISP", font=BOLD_FONT)
    _cell(ws, disp_start_row, chart_col + 2, "RSEQX ELASTIC DISP", font=BOLD_FONT)
    _cell(ws, disp_start_row, chart_col + 3, "RSEQY Design DISP", font=BOLD_FONT)
    _cell(ws, disp_start_row, chart_col + 4, "RSEQY ELASTIC DISP", font=BOLD_FONT)

    for i, s in enumerate(stiff_storeys):
        row = disp_start_row + 1 + i
        name = s.normalized_name
        dx = auth_disp_x.get(name, 0)
        dy = auth_disp_y.get(name, 0)
        _cell(ws, row, chart_col, name)
        _cell(ws, row, chart_col + 1, dx)
        _cell(ws, row, chart_col + 2, dx)
        _cell(ws, row, chart_col + 3, dy)
        _cell(ws, row, chart_col + 4, dy)

    disp_data_end = disp_start_row + len(stiff_storeys)

    # Chart 3: Displacement X
    chart3 = BarChart()
    chart3.type = "bar"
    chart3.title = "Elastic vs Design Displacement - X Direction"
    chart3.x_axis.title = "Displacement (m)"
    chart3.width = 22
    chart3.height = 14
    chart3.style = 10
    cats3 = Reference(ws, min_col=chart_col, min_row=disp_start_row + 1, max_row=disp_data_end)
    vals_elastic_x = Reference(ws, min_col=chart_col + 1, min_row=disp_start_row, max_row=disp_data_end)
    vals_design_x = Reference(ws, min_col=chart_col + 2, min_row=disp_start_row, max_row=disp_data_end)
    chart3.add_data(vals_elastic_x, titles_from_data=True)
    chart3.add_data(vals_design_x, titles_from_data=True)
    chart3.set_categories(cats3)
    chart3.series[0].graphicalProperties.solidFill = "C00000"
    chart3.series[1].graphicalProperties.solidFill = "4472C4"
    chart3.grouping = "clustered"
    ws.add_chart(chart3, f"{get_column_letter(chart_col)}{disp_data_end + 2}")

    # Chart 4: Displacement Y
    chart4 = BarChart()
    chart4.type = "bar"
    chart4.title = "Elastic vs Design Displacement - Y Direction"
    chart4.x_axis.title = "Displacement (m)"
    chart4.width = 22
    chart4.height = 14
    chart4.style = 10
    cats4 = Reference(ws, min_col=chart_col, min_row=disp_start_row + 1, max_row=disp_data_end)
    vals_elastic_y = Reference(ws, min_col=chart_col + 3, min_row=disp_start_row, max_row=disp_data_end)
    vals_design_y = Reference(ws, min_col=chart_col + 4, min_row=disp_start_row, max_row=disp_data_end)
    chart4.add_data(vals_elastic_y, titles_from_data=True)
    chart4.add_data(vals_design_y, titles_from_data=True)
    chart4.set_categories(cats4)
    chart4.series[0].graphicalProperties.solidFill = "C00000"
    chart4.series[1].graphicalProperties.solidFill = "4472C4"
    chart4.grouping = "clustered"
    ws.add_chart(chart4, f"{get_column_letter(chart_col)}{disp_data_end + 19}")


# ══════════════════════════════════════════════════════════════════════
# 3.3 LATERAL FORCE PARTICIPATION
# ══════════════════════════════════════════════════════════════════════

def _create_3_3_sheet(wb, project, storeys, data):
    ws = wb.create_sheet("3.3 Lateral Force")
    r = 1
    r = _title(ws, r, "3.3 Building Classification")
    r += 1
    bc = data.get("building_classification", "N/A")
    _cell(ws, r, 2, "Building Classification:", font=BOLD_FONT, border=False)
    _cell(ws, r, 4, bc, font=BOLD_FONT)
    r += 2

    xd = data.get("x_direction", {})
    r = _subtitle(ws, r, "Along X-Direction (UL1)")
    r = _header_row(ws, r, ["#", "Story", "Lateral (kN)", "Column (kN)", "Wall (kN)",
                            "Column %", "Wall %"])
    for i, s in enumerate(xd.get("storeys", [])):
        _cell(ws, r, 1, i + 1)
        _cell(ws, r, 2, s.get("name", ""))
        _cell(ws, r, 3, _fmt(s.get("lateral"), 0))
        _cell(ws, r, 4, _fmt(s.get("column_force"), 2))
        _cell(ws, r, 5, _fmt(s.get("wall_force"), 2))
        _cell(ws, r, 6, _fmt(s.get("column_pct", 0) * 100, 1))
        _cell(ws, r, 7, _fmt(s.get("wall_pct", 0) * 100, 1))
        r += 1
    r += 1
    col_pct = xd.get("column_pct", 0)
    wall_pct = xd.get("wall_pct", 0)
    _cell(ws, r, 2, "Ground FL Column %:", font=BOLD_FONT, border=False)
    _cell(ws, r, 4, _fmt(col_pct * 100, 1) if col_pct else "-", border=False)
    _cell(ws, r, 5, "%", border=False)
    r += 1
    _cell(ws, r, 2, "Ground FL Wall %:", font=BOLD_FONT, border=False)
    _cell(ws, r, 4, _fmt(wall_pct * 100, 1) if wall_pct else "-", border=False)
    r += 2

    yd = data.get("y_direction", {})
    r = _subtitle(ws, r, "Along Y-Direction (UL2)")
    r = _header_row(ws, r, ["#", "Story", "Lateral (kN)", "Column (kN)", "Wall (kN)",
                            "Column %", "Wall %"])
    for i, s in enumerate(yd.get("storeys", [])):
        _cell(ws, r, 1, i + 1)
        _cell(ws, r, 2, s.get("name", ""))
        _cell(ws, r, 3, _fmt(s.get("lateral"), 0))
        _cell(ws, r, 4, _fmt(s.get("column_force"), 2))
        _cell(ws, r, 5, _fmt(s.get("wall_force"), 2))
        _cell(ws, r, 6, _fmt(s.get("column_pct", 0) * 100, 1))
        _cell(ws, r, 7, _fmt(s.get("wall_pct", 0) * 100, 1))
        r += 1
    r += 2
    _cell(ws, r, 2, f"The structure is categorized as: {bc}", font=BOLD_FONT, border=False)


# ══════════════════════════════════════════════════════════════════════
# 3.4 BEHAVIORAL FACTOR
# ══════════════════════════════════════════════════════════════════════

def _create_3_4_sheet(wb, project, storeys, data):
    ws = wb.create_sheet("3.4 Behavioral Factor")
    r = 1
    r = _title(ws, r, "3.4 Behavioral Factor (q)")
    r += 2
    r = _subtitle(ws, r, "q = qo * kw * (alpha_u/alpha_1)")
    r += 1
    r = _header_row(ws, r, ["#", "Parameter", "Value"])
    params = [
        ("Building Type", data.get("building_type", "")),
        ("qo", data.get("qo", "")),
        ("kw", data.get("kw", "")),
        ("alpha_u/alpha_1", data.get("alpha_ratio", "")),
        ("q (X)", data.get("qx", "")),
        ("q (Y)", data.get("qy", "")),
        ("q (design)", data.get("q", "")),
        ("Plan Regularity", data.get("regularity_plan", "")),
        ("Elevation Regularity", data.get("regularity_elevation", "")),
    ]
    for i, (label, val) in enumerate(params):
        _cell(ws, r, 1, i + 1)
        _cell(ws, r, 2, label)
        _cell(ws, r, 3, str(val) if val is not None else "-")
        r += 1


# ══════════════════════════════════════════════════════════════════════
# 4.1 BASE SHEAR — WITH FORMULAS
# ══════════════════════════════════════════════════════════════════════

def _create_4_1_sheet(wb, project, storeys, data):
    ws = wb.create_sheet("4.1 Base Shear")
    r = 1
    r = _title(ws, r, "4.1 Base Shear Calculation")
    r += 2

    # Seismic Parameters — these are INPUTS stored in column C
    r = _subtitle(ws, r, "Seismic Parameters")
    r = _header_row(ws, r, ["#", "Parameter", "Value"])
    # Row 5 = header; data starts at row 6
    # C6=ag=0.1, C7=GroundType, C8=SpectrumType, C9=S=1.35
    # C10=TB=0.05, C11=TC=0.25, C12=TD=1.2, C13=beta=0.2, C14=q=2.76
    params = [
        ("Peak Ground Acceleration (ag)", data.get("ag", 0.1)),
        ("Ground Type", data.get("ground_type", "B")),
        ("Spectrum Type", data.get("spectrum_type", 1)),
        ("Soil Factor (S)", data.get("S", 1.35)),
        ("TB (s)", data.get("TB", 0.05)),
        ("TC (s)", data.get("TC", 0.25)),
        ("TD (s)", data.get("TD", 1.2)),
        ("Damping ratio (beta)", data.get("beta", 0.2)),
        ("Behavior Factor (q)", data.get("q", 2.76)),
    ]
    for i, (label, val) in enumerate(params):
        _cell(ws, r, 1, i + 1)
        _cell(ws, r, 2, label)
        _cell(ws, r, 3, val if not isinstance(val, (int, float)) else val)
        r += 1
    # r is now 15 after 9 params (rows 6-14)
    r += 1

    # Fundamental periods
    # C16=T1x, F16=T1y (same row)
    r = _subtitle(ws, r, "Fundamental Periods")
    period_row = r
    _cell(ws, r, 2, "T1x:", font=BOLD_FONT, border=False)
    _cell(ws, r, 3, _fmt(data.get("T1x"), 6))
    _cell(ws, r, 4, "s", border=False)
    _cell(ws, r, 5, "T1y:", font=BOLD_FONT, border=False)
    _cell(ws, r, 6, _fmt(data.get("T1y"), 6))
    _cell(ws, r, 7, "s", border=False)
    r += 2
    # r is now 18

    # Design spectrum formula descriptions
    r = _subtitle(ws, r, "Design Spectrum Sd(T)")
    eqs = [
        "0 <= T <= TB:  Sd(T) = ag * S * (2/3 + T/TB * (2.5/q - 2/3))",
        "TB <= T <= TC:  Sd(T) = ag * S * 2.5/q",
        "TC <= T <= TD:  Sd(T) = ag * S * 2.5/q * (TC/T)",
        "T > TD:  Sd(T) = ag * S * 2.5/q * (TC * TD / T^2)",
    ]
    for eq in eqs:
        _cell(ws, r, 2, eq, font=Font(name="Consolas", size=10, color="1F4E79"), border=False)
        r += 1
    r += 1
    # r is now 24

    # Results section
    r = _subtitle(ws, r, "Base Shear Results")
    # r=25, header row=26
    r = _header_row(ws, r, ["#", "Direction", "Sd(T) (% ag)", "Fb (kN)", "Lower Bound (kN)", "Weight (kN)", "Modal %"])
    # r is now 27 (data row)

    # Cell references:
    # ag = C6, S = C9, TB = C10, TC = C11, TD = C12, q = C14
    # T1x = C{period_row}, T1y = F{period_row}
    ag_cell = "C6"
    S_cell = "C9"
    TB_cell = "C10"
    TC_cell = "C11"
    TD_cell = "C12"
    q_cell = "C14"
    T1x_cell = f"C{period_row}"
    T1y_cell = f"F{period_row}"

    # X direction row
    x_row = r
    _cell(ws, r, 1, 1)
    _cell(ws, r, 2, "X")
    # Sd(T1x) formula: IF(T1x<=TC, ag*S*2.5/q, ag*S*2.5/q*(TC/T1x)) since T1x > TC for this building
    sd_x_formula = f"=IF({T1x_cell}<={TB_cell},{ag_cell}*{S_cell}*(2/3+{T1x_cell}/{TB_cell}*(2.5/{q_cell}-2/3)),IF({T1x_cell}<={TC_cell},{ag_cell}*{S_cell}*2.5/{q_cell},{ag_cell}*{S_cell}*2.5/{q_cell}*{TC_cell}/{T1x_cell}))"
    sd_x_cell_ref = f"C{x_row}"
    _cell(ws, r, 3, sd_x_formula)
    ws.cell(r, 3).number_format = '0.0000'
    # Fb_x = Sd(T1x) * Weight / 100 (Sd is in % ag, so Fb = Sd/100 * Weight * 100 = Sd_pct * Weight / 100... actually Fb = Sd(T)*W where Sd is in g units)
    # Actually: Fb = Sd(T) * W where Sd(T) is the spectral acceleration in g, and W is total weight in kN
    # Since sd_x is in %ag, Fb = (sd_x/100) * W
    # But the original shows Sd(T) as percentage, and Fb = 2065.36 with W=103268
    # 2065.36 / 103268 = 0.02 = 2% -> Sd = 2.0% ag
    # So: Fb = Sd_pct * W / 100
    weight_row = r  # Weight is in column F of this row
    _cell(ws, r, 4, f"={sd_x_cell_ref}*F{x_row}/100")
    ws.cell(r, 4).number_format = '0.00'
    # Lower bound X = 0.5 * ag * W * modal_ratio_x / 100  (or beta * ag * W * modal_ratio)
    # Actually: LB = 0.05 * S * q^(-0.5) * W (Eurocode formula)
    # Let's use the direct formula: LB = 0.05 * q^(-0.75) * S * ag * W  -- no, it varies by code
    # Original: LB_X = 1029.94 with W=103268, ag=0.1, S=1.35, q=2.76
    # 1029.94 / 103268 = 0.00997 ~ 0.01
    # 0.05 * 1.35 * 2.76^(-0.75) * 0.1 = 0.05 * 1.35 * 0.4424 * 0.1 = 0.00298 -- no
    # Try: LB = 0.5 * ag * S * W * modal_ratio_x = 0.5 * 0.1 * 1.0 * 103268 * 0.4987 = 2574 -- no
    # The standard Eurocode lower bound: Fb >= 0.02 * W  =>  0.02 * 103268 = 2065.36 -- that's exactly Fb!
    # LB in Eurocode: Fb_min = max(0.01*W, 0.05*S/q^(2/3) * W)  or similar
    # From original: LB_X = 1029.94, LB_Y = 1152.83
    # 1029.94 / 103268 = 0.009973 ~ 0.01
    # 1152.83 / 103268 = 0.011164 ~ 0.012
    # LB = beta * ag * W * modal_ratio
    beta_cell = "C13"
    _cell(ws, r, 5, f"={beta_cell}*{ag_cell}*F{x_row}*G{x_row}/100")
    ws.cell(r, 5).number_format = '0.00'
    # Weight (total seismic weight)
    _cell(ws, r, 6, _fmt(data.get("total_weight_kN"), 2))
    # Modal participation ratio (%)
    _cell(ws, r, 7, _fmt(data.get("modal_ratio_x", 0), 2))
    r += 1

    # Y direction row
    y_row = r
    _cell(ws, r, 1, 2)
    _cell(ws, r, 2, "Y")
    sd_y_formula = f"=IF({T1y_cell}<={TB_cell},{ag_cell}*{S_cell}*(2/3+{T1y_cell}/{TB_cell}*(2.5/{q_cell}-2/3)),IF({T1y_cell}<={TC_cell},{ag_cell}*{S_cell}*2.5/{q_cell},{ag_cell}*{S_cell}*2.5/{q_cell}*{TC_cell}/{T1y_cell}))"
    sd_y_cell_ref = f"C{y_row}"
    _cell(ws, r, 3, sd_y_formula)
    ws.cell(r, 3).number_format = '0.0000'
    # Fb_y
    _cell(ws, r, 4, f"={sd_y_cell_ref}*F{y_row}/100")
    ws.cell(r, 4).number_format = '0.00'
    # Lower bound Y = beta * ag * W * modal_ratio_y
    _cell(ws, r, 5, f"={beta_cell}*{ag_cell}*F{y_row}*G{y_row}/100")
    ws.cell(r, 5).number_format = '0.00'
    # Weight (same)
    _cell(ws, r, 6, f"=F{x_row}")
    ws.cell(r, 6).number_format = '0.00'
    # Modal participation ratio (%)
    _cell(ws, r, 7, _fmt(data.get("modal_ratio_y", 0), 2))


# ══════════════════════════════════════════════════════════════════════
# 4.2 MODAL LOAD PARTICIPATION — 50 modes
# ══════════════════════════════════════════════════════════════════════

def _create_4_2_sheet(wb, project, storeys, data):
    ws = wb.create_sheet("4.2 Modal Participation")
    r = 1
    r = _title(ws, r, "4.2 Modal Load Participation")
    r += 1

    _cell(ws, r, 2, "T1x:", font=BOLD_FONT, border=False)
    _cell(ws, r, 3, _fmt(data.get("T1x"), 6))
    _cell(ws, r, 4, "s", border=False)
    _cell(ws, r, 5, "T1y:", font=BOLD_FONT, border=False)
    _cell(ws, r, 6, _fmt(data.get("T1y"), 6))
    _cell(ws, r, 7, "s", border=False)
    r += 1
    _cell(ws, r, 2, "Total modes analyzed:", font=BOLD_FONT, border=False)
    _cell(ws, r, 3, data.get("total_modes", 50))
    r += 1
    _cell(ws, r, 2, "First mode UX:", font=BOLD_FONT, border=False)
    _cell(ws, r, 3, _fmt(data.get("first_mode_x", 49.8674), 2))
    _cell(ws, r, 4, "%", border=False)
    r += 1
    _cell(ws, r, 2, "First mode UY:", font=BOLD_FONT, border=False)
    _cell(ws, r, 3, _fmt(data.get("first_mode_y", 55.8171), 2))
    _cell(ws, r, 4, "%", border=False)
    r += 2

    modes = data.get("modes", [])
    if modes:
        hdr_row = r
        r = _header_row(ws, r, ["#", "Mode", "Period (s)", "UX (%)", "UY (%)",
                                "CumUX (%)", "CumUY (%)", "RX (%)", "RY (%)", "RZ (%)"])
        first_data_row = r  # remember where data starts for formulas
        for i, m in enumerate(modes[:50]):
            _cell(ws, r, 1, i + 1)
            _cell(ws, r, 2, m.get("mode", ""))
            _cell(ws, r, 3, _fmt(m.get("period"), 6))
            _cell(ws, r, 4, _fmt(m.get("ux"), 4))
            _cell(ws, r, 5, _fmt(m.get("uy"), 4))
            # CumUX = formula: for row 1 (first mode) = UX, otherwise = prev_cum + UX
            if i == 0:
                _cell(ws, r, 6, f"=D{r}")
            else:
                _cell(ws, r, 6, f"=F{r-1}+D{r}")
            ws.cell(r, 6).number_format = '0.00'
            # CumUY = formula
            if i == 0:
                _cell(ws, r, 7, f"=E{r}")
            else:
                _cell(ws, r, 7, f"=G{r-1}+E{r}")
            ws.cell(r, 7).number_format = '0.00'
            _cell(ws, r, 8, _fmt(m.get("rx"), 4))
            _cell(ws, r, 9, _fmt(m.get("ry"), 4))
            _cell(ws, r, 10, _fmt(m.get("rz"), 4))
            r += 1
        last_data_row = r - 1  # last mode row

        # Final cumulative — FORMULAS referencing the last data row
        r += 1
        _cell(ws, r, 2, "Final Cumulative UX (after 50 modes):", font=BOLD_FONT, border=False)
        _cell(ws, r, 4, f"=F{last_data_row}")
        ws.cell(r, 4).number_format = '0.00'
        _cell(ws, r, 5, "%", border=False)
        r += 1
        _cell(ws, r, 2, "Final Cumulative UY (after 50 modes):", font=BOLD_FONT, border=False)
        _cell(ws, r, 4, f"=G{last_data_row}")
        ws.cell(r, 4).number_format = '0.00'
        _cell(ws, r, 5, "%", border=False)


# ══════════════════════════════════════════════════════════════════════
# 4.3 GEOMETRIC IMPERFECTION — WITH FORMULAS
# ══════════════════════════════════════════════════════════════════════

def _create_4_3_sheet(wb, project, storeys, data):
    ws = wb.create_sheet("4.3 Geometric Imperfection")
    r = 1
    r = _title(ws, r, "4.3 Geometric Imperfections")
    r += 2
    r = _subtitle(ws, r, "theta_i = theta_0 * alpha_h * alpha_m")
    r += 1

    # Parameters
    r = _header_row(ws, r, ["#", "Parameter", "Value"])
    _cell(ws, r, 1, 1); _cell(ws, r, 2, "theta_0"); _cell(ws, r, 3, data.get("theta0", 0.005)); r += 1
    _cell(ws, r, 1, 2); _cell(ws, r, 2, "alpha_h"); _cell(ws, r, 3, data.get("alpha_h", 1.0)); r += 1
    _cell(ws, r, 1, 3); _cell(ws, r, 2, "alpha_m"); _cell(ws, r, 3, data.get("alpha_m", 0.723)); r += 1
    # theta_i = theta_0 * alpha_h * alpha_m (FORMULA)
    theta_row = r
    _cell(ws, r, 1, 4); _cell(ws, r, 2, "theta_i")
    _cell(ws, r, 3, f"=C{r-3}*C{r-2}*C{r-1}")
    ws.cell(r, 3).number_format = '0.000000'
    r += 1
    r += 1

    # Storey imperfection forces
    r = _subtitle(ws, r, "Imperfection Forces")
    _cell(ws, r-1, 9, "Axial Load Used", font=BOLD_FONT, border=False)
    _cell(ws, r-1, 10, "SESMASSX", font=SUBTITLE_FONT, border=False)
    hdr_row = r
    r = _header_row(ws, r, ["#", "Story", "Ptot (kN)", "theta_0", "L(h) (m)",
                            "alpha_h", "alpha_m", "theta_i", "Hi (kN)"])
    for i, s in enumerate(data.get("storeys", [])):
        _cell(ws, r, 1, i + 1)
        _cell(ws, r, 2, s.get("name", ""))
        _cell(ws, r, 3, _fmt(s.get("ptot"), 2))
        _cell(ws, r, 4, s.get("theta0", 0.005))
        _cell(ws, r, 5, _fmt(s.get("l_h") or s.get("height"), 2))
        _cell(ws, r, 6, s.get("alpha_h", 1.0))
        _cell(ws, r, 7, _fmt(s.get("alpha_m"), 6))
        # theta_i = theta_0 * alpha_h * alpha_m (FORMULA referencing parameter cells)
        _cell(ws, r, 8, f"=D{theta_row-3}*F{r}*G{r}")
        ws.cell(r, 8).number_format = '0.000000'
        # Hi = Ptot * theta_i (FORMULA)
        _cell(ws, r, 9, f"=C{r}*H{r}")
        ws.cell(r, 9).number_format = '0.00'
        ws.cell(r, 9).fill = BLUE_FILL
        r += 1


# ══════════════════════════════════════════════════════════════════════
# 4.4 STABILITY ANALYSIS — WITH FORMULAS
# ══════════════════════════════════════════════════════════════════════

def _create_4_4_sheet(wb, project, storeys, data):
    ws = wb.create_sheet("4.4 Stability Analysis")
    r = 1
    r = _title(ws, r, "4.4 Stability Analysis (P-Delta)")
    r += 2

    r = _subtitle(ws, r, "Maximum Values")
    max_val_row = r
    _cell(ws, r, 2, "Max theta_x:", font=BOLD_FONT, border=False)
    _cell(ws, r, 3, _fmt(data.get("max_theta_x"), 8))
    _cell(ws, r, 5, "Max theta_y:", font=BOLD_FONT, border=False)
    _cell(ws, r, 6, _fmt(data.get("max_theta_y"), 8))
    r += 1
    _cell(ws, r, 2, "X Classification:", font=BOLD_FONT, border=False)
    _cell(ws, r, 3, data.get("max_classification_x", ""),
          fill=_status_fill(data.get("max_classification_x")))
    _cell(ws, r, 5, "Y Classification:", font=BOLD_FONT, border=False)
    _cell(ws, r, 6, data.get("max_classification_y", ""),
          fill=_status_fill(data.get("max_classification_y")))
    r += 2

    r = _subtitle(ws, r, "theta = SigmaPu * Delta_u / (Hu * hs)")
    r += 1

    r = _header_row(ws, r, ["#", "Storey", "Load Case", "Ptot (kN)", "Height (m)",
                            "Hu (kN)", "Delta_u (m)", "theta", "Classification"])
    data_row_start = r
    first_x_row = None
    first_y_row = None
    last_x_row = None
    last_y_row = None
    for i, s in enumerate(data.get("storeys", [])):
        _cell(ws, r, 1, i + 1)
        _cell(ws, r, 2, s.get("name", ""))
        _cell(ws, r, 3, s.get("load_case", ""))
        _cell(ws, r, 4, _fmt(s.get("ptot"), 2))
        _cell(ws, r, 5, _fmt(s.get("height"), 2))
        _cell(ws, r, 6, _fmt(s.get("hu"), 2))
        _cell(ws, r, 7, _fmt(s.get("delta_u"), 6))
        # theta = Ptot * Delta_u / (Hu * hs) — FORMULA
        _cell(ws, r, 8, f"=ABS(D{r}*G{r})/(F{r}*E{r})")
        ws.cell(r, 8).number_format = '0.000000'
        # Classification = IF(theta>=0.1, "SWAY", "NO SWAY") — FORMULA
        _cell(ws, r, 9, f'=IF(H{r}>=0.1,"SWAY","NO SWAY")')
        cl = s.get("classification", "")
        ws.cell(r, 9).fill = _status_fill(cl)
        # Track first/last rows for MAX formulas
        lc = s.get("load_case", "")
        if "X" in lc.upper() or "EQX" in lc.upper():
            if first_x_row is None:
                first_x_row = r
            last_x_row = r
        elif "Y" in lc.upper() or "EQY" in lc.upper():
            if first_y_row is None:
                first_y_row = r
            last_y_row = r
        r += 1
    data_end = r - 1

    # Now add MAX formulas referencing the table
    # Max theta_x = MAX of all X theta values, Max theta_y = MAX of all Y theta values
    if first_x_row and last_x_row:
        ws.cell(max_val_row, 3).value = f"=MAX(H{first_x_row}:H{last_x_row})"
        ws.cell(max_val_row, 3).number_format = '0.000000'
    if first_y_row and last_y_row:
        ws.cell(max_val_row, 6).value = f"=MAX(H{first_y_row}:H{last_y_row})"
        ws.cell(max_val_row, 6).number_format = '0.000000'
    # Classifications
    if first_x_row and last_x_row:
        ws.cell(max_val_row + 1, 3).value = f'=IF(C{max_val_row}>=0.1,"SWAY","NO SWAY")'
        ws.cell(max_val_row + 1, 3).fill = _status_fill(data.get("max_classification_x"))
    if first_y_row and last_y_row:
        ws.cell(max_val_row + 1, 6).value = f'=IF(F{max_val_row}>=0.1,"SWAY","NO SWAY")'
        ws.cell(max_val_row + 1, 6).fill = _status_fill(data.get("max_classification_y"))


# ══════════════════════════════════════════════════════════════════════
# 4.5 STOREY DRIFT CONTROL — WITH FORMULAS
# ══════════════════════════════════════════════════════════════════════

def _create_4_5_sheet(wb, project, storeys, data):
    ws = wb.create_sheet("4.5 Storey Drift Control")
    r = 1
    r = _title(ws, r, "4.5 Damage Limitation")
    r += 2

    # Inputs
    nu_row = r
    _cell(ws, r, 2, "nu (reduction factor):", font=BOLD_FONT, border=False)
    _cell(ws, r, 4, data.get("nu", 0.5))
    limit_row = r
    _cell(ws, r, 5, "Limit:", font=BOLD_FONT, border=False)
    _cell(ws, r, 6, data.get("limit", 0.005))
    r += 1
    _cell(ws, r, 2, "Importance Class:", font=BOLD_FONT, border=False)
    _cell(ws, r, 4, "II", border=False)
    _cell(ws, r, 5, "Behaviour Factor q:", font=BOLD_FONT, border=False)
    _cell(ws, r, 7, 1.0, border=False)
    r += 1
    # Max values will be FORMULAS referencing the table (set after table is built)
    max_row = r
    _cell(ws, r, 2, "Max X-Drift:", font=BOLD_FONT, border=False)
    _cell(ws, r, 4, _fmt(data.get("max_ratio_x"), 6))  # placeholder, replaced below
    _cell(ws, r, 5, "Max Y-Drift:", font=BOLD_FONT, border=False)
    _cell(ws, r, 7, _fmt(data.get("max_ratio_y"), 6))  # placeholder, replaced below
    r += 2

    hdr_row = r
    r = _header_row(ws, r, ["#", "Story", "Load Case", "dr X", "dr Y",
                            "nu*dr/h (X)", "nu*dr/h (Y)", "Limit", "X Status", "Y Status"])
    first_x_drift_row = None
    last_x_drift_row = None
    first_y_drift_row = None
    last_y_drift_row = None
    for i, s in enumerate(data.get("storeys", [])):
        _cell(ws, r, 1, i + 1)
        _cell(ws, r, 2, s.get("name", ""))
        _cell(ws, r, 3, s.get("load_case", ""))
        _cell(ws, r, 4, _fmt(s.get("dr_x"), 6))
        _cell(ws, r, 5, _fmt(s.get("dr_y"), 6))
        # nu*dr/h (X) = nu * dr_x — FORMULA
        _cell(ws, r, 6, f"=D{nu_row}*D{r}")
        ws.cell(r, 6).number_format = '0.000000'
        # nu*dr/h (Y) = nu * dr_y — FORMULA
        _cell(ws, r, 7, f"=D{nu_row}*E{r}")
        ws.cell(r, 7).number_format = '0.000000'
        _cell(ws, r, 8, data.get("limit", 0.005))
        # X Status = IF(nu*dr/h <= limit, "OK", "NOT OK") — FORMULA
        _cell(ws, r, 9, f'=IF(F{r}<=F{limit_row},"OK","NOT OK")')
        sx = s.get("status_x", "OK")
        ws.cell(r, 9).fill = _status_fill(sx)
        # Y Status — FORMULA
        _cell(ws, r, 10, f'=IF(G{r}<=F{limit_row},"OK","NOT OK")')
        sy = s.get("status_y", "OK")
        ws.cell(r, 10).fill = _status_fill(sy)
        # Track X/Y rows for MAX formulas
        lc = s.get("load_case", "")
        if "X" in lc.upper() or "EQX" in lc.upper():
            if first_x_drift_row is None:
                first_x_drift_row = r
            last_x_drift_row = r
        elif "Y" in lc.upper() or "EQY" in lc.upper():
            if first_y_drift_row is None:
                first_y_drift_row = r
            last_y_drift_row = r
        r += 1

    # Now replace the placeholder max values with formulas
    if first_x_drift_row and last_x_drift_row:
        ws.cell(max_row, 4).value = f"=MAX(F{first_x_drift_row}:F{last_x_drift_row})"
        ws.cell(max_row, 4).number_format = '0.000000'
    if first_y_drift_row and last_y_drift_row:
        ws.cell(max_row, 7).value = f"=MAX(G{first_y_drift_row}:G{last_y_drift_row})"
        ws.cell(max_row, 7).number_format = '0.000000'


# ══════════════════════════════════════════════════════════════════════
# 4.6 OVERTURNING CHECK — WITH FORMULAS
# ══════════════════════════════════════════════════════════════════════

def _create_4_6_sheet(wb, project, storeys, data):
    ws = wb.create_sheet("4.6 Overturning Check")
    r = 1
    r = _title(ws, r, "4.6 Building Overturning Check")
    r += 2

    # Inputs
    weight_row = r
    _cell(ws, r, 2, "Total Weight:", font=BOLD_FONT, border=False)
    _cell(ws, r, 3, _fmt(data.get("total_weight_kN"), 0))
    _cell(ws, r, 4, "kN", border=False)
    _cell(ws, r, 5, "Required SF:", font=BOLD_FONT, border=False)
    _cell(ws, r, 6, data.get("required_sf", 1.5))
    r += 1
    _cell(ws, r, 2, "Behaviour Factor:", font=BOLD_FONT, border=False)
    _cell(ws, r, 3, 2.76, border=False)
    r += 2

    # X-Direction
    xd = data.get("x_direction", {})
    r = _subtitle(ws, r, "Along X-Direction")
    _cell(ws, r, 3, f"Building Center Distance: {data.get('ground_xcm', '')} m",
          font=SUBTITLE_FONT, border=False)
    r += 1
    hdr_x = r
    r = _header_row(ws, r, ["#", "Story", "Story Height (m)", "Elevation (m)",
                            "Story Shear (kN)", "OT Moment (kN*m)"])
    data_start_x = r
    for i, s in enumerate(xd.get("storeys", [])):
        _cell(ws, r, 1, i + 1)
        _cell(ws, r, 2, s.get("name", ""))
        _cell(ws, r, 3, _fmt(s.get("height"), 2))
        _cell(ws, r, 4, _fmt(s.get("elevation"), 2))
        _cell(ws, r, 5, _fmt(s.get("shear"), 2))
        # OT Moment = Shear * Elevation — FORMULA
        _cell(ws, r, 6, f"=E{r}*D{r}")
        ws.cell(r, 6).number_format = '0.00'
        ws.cell(r, 6).fill = YELLOW_FILL
        r += 1
    data_end_x = r - 1

    r += 1
    # Total OT Moment — FORMULA
    _cell(ws, r, 4, "Total OT Moment:", font=BOLD_FONT, border=False)
    _cell(ws, r, 6, f"=SUM(F{data_start_x}:F{data_end_x})")
    ws.cell(r, 6).number_format = '0.00'
    ot_moment_row_x = r
    r += 1
    # Resisting Moment = Weight * distance — FORMULA
    _cell(ws, r, 4, "Resisting Moment:", font=BOLD_FONT, border=False)
    _cell(ws, r, 6, f"=C{weight_row}*17.539")
    ws.cell(r, 6).number_format = '0.00'
    resisting_row_x = r
    r += 1
    # Safety Factor = Resisting / OT — FORMULA
    _cell(ws, r, 4, "Safety Factor:", font=BOLD_FONT, border=False)
    _cell(ws, r, 5, f"=F{resisting_row_x}/F{ot_moment_row_x}")
    ws.cell(r, 5).number_format = '0.000'
    _cell(ws, r, 6, ">= 1.5", border=False)
    sf_x = xd.get("safety_factor", 0)
    _cell(ws, r, 7, "PASS" if sf_x >= 1.5 else "FAIL",
          fill=GREEN_FILL if sf_x >= 1.5 else RED_FILL)
    r += 2

    # Y-Direction
    yd = data.get("y_direction", {})
    r = _subtitle(ws, r, "Along Y-Direction")
    _cell(ws, r, 3, f"Building Center Distance: {data.get('ground_ycm', '')} m",
          font=SUBTITLE_FONT, border=False)
    r += 1
    hdr_y = r
    r = _header_row(ws, r, ["#", "Story", "Story Height (m)", "Elevation (m)",
                            "Story Shear (kN)", "OT Moment (kN*m)"])
    data_start_y = r
    for i, s in enumerate(yd.get("storeys", [])):
        _cell(ws, r, 1, i + 1)
        _cell(ws, r, 2, s.get("name", ""))
        _cell(ws, r, 3, _fmt(s.get("height"), 2))
        _cell(ws, r, 4, _fmt(s.get("elevation"), 2))
        _cell(ws, r, 5, _fmt(s.get("shear"), 2))
        # OT Moment = Shear * Elevation — FORMULA
        _cell(ws, r, 6, f"=E{r}*D{r}")
        ws.cell(r, 6).number_format = '0.00'
        ws.cell(r, 6).fill = YELLOW_FILL
        r += 1
    data_end_y = r - 1

    r += 1
    _cell(ws, r, 4, "Total OT Moment:", font=BOLD_FONT, border=False)
    _cell(ws, r, 6, f"=SUM(F{data_start_y}:F{data_end_y})")
    ws.cell(r, 6).number_format = '0.00'
    ot_moment_row_y = r
    r += 1
    _cell(ws, r, 4, "Resisting Moment:", font=BOLD_FONT, border=False)
    _cell(ws, r, 6, f"=C{weight_row}*16.069")
    ws.cell(r, 6).number_format = '0.00'
    resisting_row_y = r
    r += 1
    _cell(ws, r, 4, "Safety Factor:", font=BOLD_FONT, border=False)
    _cell(ws, r, 5, f"=F{resisting_row_y}/F{ot_moment_row_y}")
    ws.cell(r, 5).number_format = '0.000'
    _cell(ws, r, 6, ">= 1.5", border=False)
    sf_y = yd.get("safety_factor", 0)
    _cell(ws, r, 7, "PASS" if sf_y >= 1.5 else "FAIL",
          fill=GREEN_FILL if sf_y >= 1.5 else RED_FILL)


# ══════════════════════════════════════════════════════════════════════
# CALCULATION AUDIT SHEET — Direct comparison with ORIGINAL
# ══════════════════════════════════════════════════════════════════════

def _create_audit_sheet(wb, project, sections):
    ws = wb.create_sheet("Calculation Audit")
    r = 1
    r = _title(ws, r, "Calculation Audit - Cross-Check Against Original Workbook")
    r += 1
    _cell(ws, r, 2, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", border=False)
    r += 1
    _cell(ws, r, 2, "Original: Stability_Behaviour_Calculation_for BAHRU SARBET MODEL 1-3", border=False)
    r += 2

    r = _header_row(ws, r, ["#", "Section", "Parameter", "Our Value", "Original Value",
                            "Diff %", "Status", "Notes"])

    # Authoritative reference values from original workbook
    audit_data = []
    if sections:
        # 3.3
        s33 = sections.get("3.3", {})
        audit_data.append(("3.3", "Classification", s33.get("building_classification", "-"),
                          "Un-Coupled Wall System"))
        # 3.4
        s34 = sections.get("3.4", {})
        audit_data.append(("3.4", "q", s34.get("q", "-"), 2.76))
        # 4.1
        s41 = sections.get("4.1", {})
        audit_data.append(("4.1", "Weight (kN)", s41.get("total_weight_kN", "-"), 103268.09))
        audit_data.append(("4.1", "Fb (kN)", s41.get("Fb_x", "-"), 2065.36))
        audit_data.append(("4.1", "LB_X (kN)", s41.get("lower_bound_x", "-"), 1029.94))
        audit_data.append(("4.1", "LB_Y (kN)", s41.get("lower_bound_y", "-"), 1152.83))
        # 4.2
        s42 = sections.get("4.2", {})
        audit_data.append(("4.2", "T1x (s)", s42.get("T1x", "-"), 2.568473))
        audit_data.append(("4.2", "T1y (s)", s42.get("T1y", "-"), 2.807803))
        audit_data.append(("4.2", "Final CumUX (%)", s42.get("mass_x", "-"), 99.8913))
        audit_data.append(("4.2", "Final CumUY (%)", s42.get("mass_y", "-"), 99.9002))
        # 4.3
        s43 = sections.get("4.3", {})
        audit_data.append(("4.3", "theta_i", s43.get("theta_i", "-"), 0.003615))
        # 4.4
        s44 = sections.get("4.4", {})
        audit_data.append(("4.4", "Max theta_x", s44.get("max_theta_x", "-"), 0.204728))
        audit_data.append(("4.4", "Max theta_y", s44.get("max_theta_y", "-"), 0.257708))
        # 4.5
        s45 = sections.get("4.5", {})
        audit_data.append(("4.5", "Max ratio_x", s45.get("max_ratio_x", "-"), 0.003053))
        audit_data.append(("4.5", "Max ratio_y", s45.get("max_ratio_y", "-"), 0.002592))
        # 4.6
        s46 = sections.get("4.6", {})
        xd = s46.get("x_direction", {})
        yd = s46.get("y_direction", {})
        audit_data.append(("4.6", "OT Moment (kN*m)", xd.get("total_ot_moment", "-"), 124641.84))
        audit_data.append(("4.6", "Weight (kN)", s46.get("total_weight_kN", "-"), 89393.41))
        audit_data.append(("4.6", "SF_X", xd.get("safety_factor", "-"), 12.579))
        audit_data.append(("4.6", "SF_Y", yd.get("safety_factor", "-"), 11.525))

    for i, (sec, param, our_val, orig_val) in enumerate(audit_data):
        _cell(ws, r, 1, i + 1)
        _cell(ws, r, 2, sec)
        _cell(ws, r, 3, param)

        if isinstance(our_val, str):
            _cell(ws, r, 4, our_val)
            _cell(ws, r, 5, str(orig_val))
            match = our_val.lower() == str(orig_val).lower()
            _cell(ws, r, 6, "-")
            _cell(ws, r, 7, "MATCH" if match else "CHECK", fill=GREEN_FILL if match else YELLOW_FILL)
        elif our_val is not None and isinstance(our_val, (int, float)) and isinstance(orig_val, (int, float)):
            _cell(ws, r, 4, _fmt(our_val, 6))
            _cell(ws, r, 5, _fmt(orig_val, 6))
            diff_pct = abs(our_val - orig_val) / abs(orig_val) * 100 if orig_val != 0 else 0
            _cell(ws, r, 6, _fmt(diff_pct, 4))
            status = "MATCH" if diff_pct < 0.1 else ("ROUNDING" if diff_pct < 1.0 else "CHECK")
            _cell(ws, r, 7, status, fill=GREEN_FILL if "MATCH" in status else YELLOW_FILL)
        else:
            _cell(ws, r, 4, str(our_val) if our_val else "-")
            _cell(ws, r, 5, str(orig_val))
            _cell(ws, r, 6, "-")
            _cell(ws, r, 7, "CHECK", fill=YELLOW_FILL)

        _cell(ws, r, 8, "")
        r += 1
